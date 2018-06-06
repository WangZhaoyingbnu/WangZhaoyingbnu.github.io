#coding=utf-8
import string
import os
import time
import types
import pdb
from datetime import datetime, timedelta
from openpyxl import load_workbook
#wb = Workbook()
#ws = wb.active

#root directory

#ffmpeg_path = r'C:/Users/lidongxue/Desktop/Ann_Wang/Downloads/ffmpeg/bin'
#ffmpeg_name = ffmpeg_path + '/ffmpeg'

#video_path = r'F:'
#video_file_path = r'F:/pre-process data.xlsx'
wb = load_workbook(filename = r'F:/pre_process_data.xlsx')
ws = wb.worksheets[0]

ff_input_hsv = r'F:/HSV'
ff_input_hdv = r'F:/HDV'

#ff_output_hsv = r'C:/Users/lidongxue/Desktop/Ann-Wang/Thesis/processed data/hsv'
ff_output = r' C:/Users/lidongxue/Desktop/Ann_Wang/Thesis/processed_data'
ff_output_hsv = ff_output + '/hsv'
ff_output_hdv = ff_output + '/hdv'


#standard time
standard_time = datetime.strptime('0:0:1','%H:%M:%S')

def time_input(a):
	if isinstance(a,str):
		return datetime.strptime(a,'%H:%M:%S')
	else:
		return a

#transfer videos
row_actual = range(105,451)
for row in row_actual:
	if ws.cell(row = row, column = 2).value == 1:
		if ws.cell(row = row, column = 13).value == 1:
			hsv_name = ws.cell(row, 7).value
			hdv_name = ws.cell(row, 10).value

			hsv_path = ff_input_hsv + '/' + hsv_name + '.MXF'
			hdv_path = ff_input_hdv + '/' + hdv_name + '/' + hdv_name + '.MP4'
			#hdv_path = r'F:/HDV/153_0006_01/153_0006_01.MP4'

			hsv_output = ff_output_hsv + '/' + str(row) +'.MXF'
			hdv_output = ff_output_hdv + '/' + str(row) +'.MP4'

			#hsvs = datetime.strptime(ws.cell(row,8).value,'%H:%M:%S')
			#hsve = datetime.strptime(ws.cell(row,9).value,'%H:%M:%S')
			#hsvs = ws.cell(row,8).value
			#hsve = ws.cell(row,9).value
			#hsvs = hsvs.time()
			hsvs = time_input(ws.cell(row,8).value)
			hsve = time_input(ws.cell(row,9).value)

			#hdvs = datetime.strptime(ws.cell(row,11).value,'%H:%M:%S')
			#hdve = datetime.strptime(ws.cell(row,12).value,'%H:%M:%S')
			#hdvs = ws.cell(row,11).value
			#hdve = ws.cell(row,12).value
			#hdvs = hdvs.time()
			hdvs = time_input(ws.cell(row,11).value)
			hdve = time_input(ws.cell(row,12).value)

			#duration_hsv = standard_time + hsve - hsvs
			#duration_hdv = standard_time + hdve - hdvs
			
			duration_hsv = timedelta(minutes = hsve.minute, seconds = hsve.second) - timedelta(minutes = hsvs.minute, seconds = hsvs.second) + standard_time
			duration_hdv = timedelta(minutes = hdve.minute, seconds = hdve.second) - timedelta(minutes = hdvs.minute, seconds = hdvs.second) + standard_time
			duration_hsv = duration_hsv.time()
			duration_hdv = duration_hdv.time()

			#os.system(ffmpeg -ss 0:0:0 -t 0:0:2 -i )
			hdv_str = 'ffmpeg -i -ss %s -t %s '% (hdvs, duration_hdv)
			hsv_str = 'ffmpeg -i %s -ss %s -t %s '% (hsv_path, hsvs, duration_hsv)
			os.system( hdv_str + hdv_path + hdv_output )
			os.system( hsv_str + hsv_output )


