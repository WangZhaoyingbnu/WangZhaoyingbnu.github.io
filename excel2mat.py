import scipy.io as scio
import numpy as np
from openpyxl import load_workbook

normalizor = [30., 10000., 200., 200.]
wb = load_workbook(filename = r'F:/pre_process_data.xlsx')
ws = wb.worksheets[0]
row = ws.max_row
array = np.zeros((row, 4)) #[m,n]

for i in range(1, row + 1):
	for j in range(3, 7):
		array[i-1, j-3] = ws.cell(row = i, column = j).value
		array[i-1, j-3] = array[i-1, j-3]/normalizor[j-3]



scio.savemat('C:/Users/lidongxue/Desktop/Ann_Wang/Thesis/processed_data1/parameter.mat',{"Parameter" : array})