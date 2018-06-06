#coding=utf-8
import string
import os
import time
import types
import pdb
import math
from datetime import datetime, timedelta
from openpyxl import load_workbook

#root directory
#ffmpeg_name = r'C:/Users/lidongxue/Desktop/Ann_Wang/Downloads/ffmpeg/bin/ffmpeg'
ffmpeg_path = r'C:/Users/lidongxue/Desktop/Ann_Wang/Downloads/ffmpeg/bin'
ffmpeg_name = ffmpeg_path + '/ffmpeg'

wb = load_workbook(filename = r'F:/pre_process_data.xlsx')

ws = wb.worksheets[0]
rows = ws.rows
columns = ws.columns

ff_input_hsv = r'F:/HSV'
ff_input_hdv = r'F:/HDV'

#ff_output_hsv = r'C:/Users/lidongxue/Desktop/Ann-Wang/Thesis/processed data/hsv'
ff_output = r'C:/Users/lidongxue/Desktop/Ann_Wang/Thesis/processed_data1'
ff_output_hsv = ff_output + '/hsv'
ff_output_hdv = ff_output + '/hdv'
######################################################
ff_output_picture = ff_output + '/picture_hdv'
#ff_output_picture = ff_output + '/picture_hsv'

standard_time = datetime.strptime('0:0:1','%H:%M:%S')

def time_input(a):
	if isinstance(a, str):
		return datetime.strptime(a,'%H:%M:%S')
	else:
		return a

def time_tran(b):
	if not isinstance(b, str):
		return '0:%d:%d' %(b.minute, b.second)

def ensure_dir(f):
	if not os.path.exists(f):
		os.makedirs(f)

#transfer videos
for row in range(1,451):
	if ws.cell(row = row, column = 2).value == 1 or ws.cell(row = row, column = 2).value == -1:
		if ws.cell(row = row, column = 13).value == 1:
			picture_output_1 = ff_output_picture + '/' + str(row)
			ensure_dir(picture_output_1)
			def picture_cap(num_picture):
				return picture_output_1 + '/' + str(num_picture) + '.JPEG'

			#1:hsvs, 2:hsve, 3:hdvs, 4:hdve
			#
			'''
			hsv_name = ws.cell(row, 7).value
			hsv_path = ff_input_hsv + '/' + hsv_name + '.MXF'
			hsv_output = ff_output_hsv + '/' + str(row) + '.MXF'
			picture_output_hsv = list(map(picture_cap, range(1,3)))
			hsvs = time_input(ws.cell(row,8).value)
			hsve = time_input(ws.cell(row,9).value)
			hsve = timedelta(minutes = hsve.minute, seconds = hsve.second) + standard_time
			htime_hsv = [hsvs, hsve]
			htime_hsv = list(map(time_tran, htime_hsv))
			h_path_hsv = [hsv_path, hsv_path]
			for k in range(2):
				os.system('ffmpeg -ss %(time)s -i %(path)s -frames:v 1 %(path2)s ' %{'time':htime_hsv[k],'path':h_path_hsv[k],'path2':picture_output_hsv[k]})
			'''

			hdv_name = ws.cell(row, 10).value
			hdv_path = ff_input_hdv + '/' + hdv_name + '/' + hdv_name + '.MP4'
			hdv_output = ff_output_hdv + '/' + str(row) + '.MP4'
			picture_output_hdv = list(map(picture_cap, range(3,5)))
			hdvs = time_input(ws.cell(row,11).value)
			hdve = time_input(ws.cell(row,12).value)
			hdve = timedelta(minutes = hdve.minute, seconds = hdve.second) + standard_time
			htime_hdv = [hdvs, hdve]
			htime_hdv = list(map(time_tran, htime_hdv))
			h_path_hdv = [hdv_path, hdv_path]
			for k in range(2):
				os.system('ffmpeg -ss %(time)s -i %(path)s -frames:v 1 %(path2)s ' %{'time':htime_hdv[k],'path':h_path_hdv[k],'path2':picture_output_hdv[k]})
			#os.system('ffmpeg -ss 0:00:00 -i F:/HSV/000_0006.MXF -frames:v 1 C:/Users/lidongxue/Desktop/Ann_Wang/Thesis/processed_data/1/1.JPEG')
			


